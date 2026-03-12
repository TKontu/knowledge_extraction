"""Tests for consolidated table report generation (unified 3-sheet mode)."""

from unittest.mock import MagicMock

import pytest

from services.reports.consolidated_builder import (
    ConsolidatedReportBuilder,
    SheetData,
    build_provenance_sheets,
    render_markdown,
)
from services.reports.excel_formatter import ExcelFormatter
from services.reports.schema_table_generator import SchemaTableGenerator


# ── Fixtures ──


@pytest.fixture
def sample_schema():
    """Schema with 2 scalar groups and 1 entity list."""
    return {
        "name": "drivetrain_schema",
        "extraction_context": {"source_label": "Company"},
        "field_groups": [
            {
                "name": "company_meta",
                "description": "Company metadata",
                "prompt_hint": "Extract company info",
                "fields": [
                    {"name": "company_name", "field_type": "text", "description": "Name"},
                    {"name": "founded_year", "field_type": "integer", "description": "Founded year"},
                    {"name": "is_oem", "field_type": "boolean", "description": "Is OEM"},
                ],
            },
            {
                "name": "certifications",
                "description": "Company certifications",
                "prompt_hint": "Extract certs",
                "fields": [
                    {"name": "iso_certified", "field_type": "boolean", "description": "ISO certified"},
                    {"name": "cert_list", "field_type": "list", "description": "Certifications"},
                ],
            },
            {
                "name": "products_gearbox",
                "description": "Gearbox products",
                "prompt_hint": "Extract products",
                "is_entity_list": True,
                "fields": [
                    {"name": "product_name", "field_type": "text", "description": "Product"},
                    {"name": "max_torque_nm", "field_type": "float", "description": "Max torque"},
                    {"name": "gear_type", "field_type": "enum", "description": "Gear type",
                     "enum_values": ["spur", "helical", "planetary"]},
                ],
            },
        ],
    }


@pytest.fixture
def collision_schema():
    """Schema with field name collisions across groups."""
    return {
        "name": "collision_schema",
        "field_groups": [
            {
                "name": "group_a",
                "description": "Group A",
                "prompt_hint": "...",
                "fields": [
                    {"name": "name", "field_type": "text", "description": "Name A"},
                    {"name": "unique_a", "field_type": "text", "description": "Unique A"},
                ],
            },
            {
                "name": "group_b",
                "description": "Group B",
                "prompt_hint": "...",
                "fields": [
                    {"name": "name", "field_type": "text", "description": "Name B"},
                    {"name": "unique_b", "field_type": "text", "description": "Unique B"},
                ],
            },
        ],
    }


@pytest.fixture
def generator():
    return SchemaTableGenerator()


def _make_record(source_group, extraction_type, data, provenance=None, source_count=3):
    """Create a mock ConsolidatedExtraction record."""
    rec = MagicMock()
    rec.source_group = source_group
    rec.extraction_type = extraction_type
    rec.data = data
    rec.provenance = provenance or {}
    rec.source_count = source_count
    rec.grounded_count = 2
    return rec


@pytest.fixture
def sample_records():
    """Two companies with scalar + entity data."""
    return [
        # Company A - scalar
        _make_record("Acme Corp", "company_meta", {
            "company_name": "Acme Corporation",
            "founded_year": 1995,
            "is_oem": True,
        }, provenance={
            "company_name": {"winning_weight": 0.9, "top_sources": ["s1", "s2"]},
            "founded_year": {"winning_weight": 0.8, "top_sources": ["s1"]},
            "is_oem": {"winning_weight": 1.0, "top_sources": ["s1"]},
        }),
        _make_record("Acme Corp", "certifications", {
            "iso_certified": True,
            "cert_list": ["ISO 9001", "ISO 14001"],
        }),
        # Company A - entity
        _make_record("Acme Corp", "products_gearbox", {
            "products_gearbox": [
                {"product_name": "GearX-100", "max_torque_nm": 500.0, "gear_type": "helical"},
                {"product_name": "GearX-200", "max_torque_nm": 1000.0, "gear_type": "planetary"},
            ]
        }, provenance={
            "products_gearbox": {
                "winning_weight": 0.85,
                "top_sources": ["s1"],
                "entity_provenance": [
                    {"winning_weight": 0.9, "top_sources": ["s1"]},
                    {"winning_weight": 0.8, "top_sources": ["s1"]},
                ],
            },
        }),
        # Company B - scalar
        _make_record("Beta Inc", "company_meta", {
            "company_name": "Beta Industries",
            "founded_year": 2003,
            "is_oem": False,
        }),
        _make_record("Beta Inc", "certifications", {
            "iso_certified": False,
            "cert_list": ["CE"],
        }),
        # Company B - entity
        _make_record("Beta Inc", "products_gearbox", {
            "products_gearbox": [
                {"product_name": "BetaDrive", "max_torque_nm": 250.0, "gear_type": "spur"},
            ]
        }),
    ]


# ── TestGetScalarColumns ──


class TestGetScalarColumns:
    def test_excludes_entity_list(self, generator, sample_schema):
        columns, labels, col_types = generator.get_scalar_columns(sample_schema)
        assert "products_gearbox" not in columns
        assert "product_name" not in columns

    def test_includes_source_group(self, generator, sample_schema):
        columns, labels, _ = generator.get_scalar_columns(sample_schema)
        assert columns[0] == "source_group"
        assert labels["source_group"] == "Source"

    def test_no_source_url_columns(self, generator, sample_schema):
        columns, _, _ = generator.get_scalar_columns(sample_schema)
        assert "source_url" not in columns
        assert "source_title" not in columns
        assert "domain" not in columns
        assert "avg_confidence" not in columns

    def test_collision_prefixing(self, generator, collision_schema):
        columns, labels, _ = generator.get_scalar_columns(collision_schema)
        assert "group_a.name" in columns
        assert "group_b.name" in columns
        assert "unique_a" in columns
        assert "unique_b" in columns


# ── TestGetEntityGroupColumns ──


class TestGetEntityGroupColumns:
    def test_all_fields_for_group(self, generator, sample_schema):
        columns, labels, col_types = generator.get_entity_group_columns(
            sample_schema, "products_gearbox"
        )
        assert "source_group" in columns
        assert "product_name" in columns
        assert "max_torque_nm" in columns
        assert "gear_type" in columns

    def test_source_group_prepended(self, generator, sample_schema):
        columns, _, _ = generator.get_entity_group_columns(
            sample_schema, "products_gearbox"
        )
        assert columns[0] == "source_group"

    def test_unit_inference_in_labels(self, generator, sample_schema):
        _, labels, _ = generator.get_entity_group_columns(
            sample_schema, "products_gearbox"
        )
        assert "Nm" in labels["max_torque_nm"]

    def test_invalid_group_raises(self, generator, sample_schema):
        with pytest.raises(ValueError, match="not found"):
            generator.get_entity_group_columns(sample_schema, "nonexistent")


# ── TestGetUnifiedColumns ──


class TestGetUnifiedColumns:
    def test_includes_scalars_and_entity_list(self, generator, sample_schema):
        columns, labels, col_types, entity_groups = generator.get_unified_columns(sample_schema)
        # Scalar fields present
        assert "company_name" in columns
        assert "founded_year" in columns
        assert "is_oem" in columns
        assert "iso_certified" in columns
        assert "cert_list" in columns
        # Entity list as summary column
        assert "products_gearbox_list" in columns
        assert col_types["products_gearbox_list"] == "entity_list"
        assert "products_gearbox_list" in entity_groups

    def test_source_label_from_context(self, generator, sample_schema):
        columns, labels, _, _ = generator.get_unified_columns(sample_schema)
        assert labels["source_group"] == "Company"

    def test_source_label_fallback(self, generator):
        schema = {
            "name": "test",
            "field_groups": [
                {"name": "g", "description": "G", "prompt_hint": "...",
                 "fields": [{"name": "x", "field_type": "text", "description": "X"}]},
            ],
        }
        _, labels, _, _ = generator.get_unified_columns(schema)
        assert labels["source_group"] == "Source"

    def test_collision_prefixing(self, generator, collision_schema):
        columns, _, _, _ = generator.get_unified_columns(collision_schema)
        assert "group_a.name" in columns
        assert "group_b.name" in columns


# ── TestBuildUnifiedRow ──


class TestBuildUnifiedRow:
    def test_merges_scalar_groups(self, generator, sample_schema, sample_records):
        builder = ConsolidatedReportBuilder(generator)
        data_sheet, summary = builder.gather(sample_records, sample_schema)
        acme_row = data_sheet.rows[0]
        assert acme_row["source_group"] == "Acme Corp"
        assert acme_row["company_name"] == "Acme Corporation"
        assert acme_row["founded_year"] == 1995
        assert acme_row["is_oem"] is True

    def test_flat_list_inline(self, generator, sample_schema, sample_records):
        builder = ConsolidatedReportBuilder(generator)
        data_sheet, _ = builder.gather(sample_records, sample_schema)
        acme_row = data_sheet.rows[0]
        assert acme_row["cert_list"] == "ISO 9001, ISO 14001"

    def test_entity_list_formatted_cell(self, generator, sample_schema, sample_records):
        builder = ConsolidatedReportBuilder(generator)
        data_sheet, _ = builder.gather(sample_records, sample_schema)
        acme_row = data_sheet.rows[0]
        # Entity list should be formatted string with all fields
        entity_cell = acme_row["products_gearbox_list"]
        assert "GearX-100" in entity_cell
        assert "GearX-200" in entity_cell
        assert "500.0Nm" in entity_cell
        # New: all fields shown, including gear_type
        assert "helical" in entity_cell
        assert "planetary" in entity_cell
        # Items separated by newlines
        assert "\n" in entity_cell

    def test_entity_list_filters_low_quality(self, generator, sample_schema):
        """Entities with low winning_weight are excluded from cell."""
        records = [
            _make_record("FilterCorp", "company_meta", {"company_name": "FilterCorp"}),
            _make_record("FilterCorp", "products_gearbox", {
                "products_gearbox": [
                    {"product_name": "Good", "max_torque_nm": 100.0, "gear_type": "spur"},
                    {"product_name": "Bad", "max_torque_nm": 50.0, "gear_type": "worm"},
                ]
            }, provenance={
                "products_gearbox": {
                    "winning_weight": 0.5,
                    "top_sources": ["s1"],
                    "entity_provenance": [
                        {"winning_weight": 0.9, "top_sources": ["s1"]},  # Good — above 0.3
                        {"winning_weight": 0.1, "top_sources": ["s1"]},  # Bad — below 0.3
                    ],
                },
            }),
        ]
        builder = ConsolidatedReportBuilder(generator)
        data_sheet, _ = builder.gather(records, sample_schema)
        cell = data_sheet.rows[0]["products_gearbox_list"]
        assert "Good" in cell
        assert "Bad" not in cell

    def test_list_of_dicts_formatted(self, generator, sample_schema):
        """List-of-dicts scalar values are formatted as delimited items, not 'N items'."""
        records = [
            _make_record("LocCorp", "company_meta", {
                "company_name": "LocCorp",
                "founded_year": 2000,
                "is_oem": False,
            }),
            _make_record("LocCorp", "certifications", {
                "iso_certified": True,
                "cert_list": [
                    {"city": "Munich", "country": "Germany", "site_type": "HQ"},
                    {"city": "Tokyo", "country": "Japan", "site_type": "Branch"},
                ],
            }),
        ]
        builder = ConsolidatedReportBuilder(generator)
        data_sheet, _ = builder.gather(records, sample_schema)
        cell = data_sheet.rows[0]["cert_list"]
        assert "Munich" in cell
        assert "Germany" in cell
        assert "Tokyo" in cell
        assert "items" not in cell  # Not "2 items"

    def test_empty_entity_list_na(self, generator, sample_schema):
        records = [
            _make_record("Empty Corp", "company_meta", {"company_name": "Empty"}),
            _make_record("Empty Corp", "products_gearbox", {"products_gearbox": []}),
        ]
        builder = ConsolidatedReportBuilder(generator)
        data_sheet, _ = builder.gather(records, sample_schema)
        row = data_sheet.rows[0]
        assert row["products_gearbox_list"] == "N/A"

    def test_missing_extraction_type_none(self, generator, sample_schema):
        records = [
            _make_record("Lonely Corp", "company_meta", {
                "company_name": "Lonely",
                "founded_year": 2020,
                "is_oem": False,
            }),
            # No certifications record
        ]
        builder = ConsolidatedReportBuilder(generator)
        data_sheet, _ = builder.gather(records, sample_schema)
        row = data_sheet.rows[0]
        assert row.get("iso_certified") is None

    def test_gather_returns_tuple(self, generator, sample_schema, sample_records):
        builder = ConsolidatedReportBuilder(generator)
        result = builder.gather(sample_records, sample_schema)
        assert isinstance(result, tuple)
        assert len(result) == 2
        data_sheet, summary = result
        assert isinstance(data_sheet, SheetData)
        assert isinstance(summary, dict)
        assert "total_count" in summary

    def test_summary_counts(self, generator, sample_schema, sample_records):
        builder = ConsolidatedReportBuilder(generator)
        _, summary = builder.gather(sample_records, sample_schema)
        assert summary["total_count"] == 2
        assert "Products Gearbox" in summary["entity_counts"]
        # Acme has 2 + Beta has 1 = 3
        assert summary["entity_counts"]["Products Gearbox"] == 3

    def test_sheet_name_uses_source_label(self, generator, sample_schema, sample_records):
        builder = ConsolidatedReportBuilder(generator)
        data_sheet, _ = builder.gather(sample_records, sample_schema)
        assert data_sheet.name == "Company Data"


# ── TestEntityPagination ──


class TestEntityPagination:
    def test_no_pagination_when_under_page_size(self, generator, sample_schema, sample_records):
        """No extra columns when entity count < page_size."""
        builder = ConsolidatedReportBuilder(generator)
        data_sheet, _ = builder.gather(sample_records, sample_schema, page_size=50)
        # Only one products_gearbox_list column, no _p2
        entity_cols = [c for c in data_sheet.columns if "products_gearbox" in c]
        assert entity_cols == ["products_gearbox_list"]

    def test_pagination_creates_page_columns(self, generator, sample_schema):
        """Entity list with more items than page_size creates page columns."""
        # Create 5 products — with page_size=2, should get 3 page columns
        products = [
            {"product_name": f"Gear{i}", "max_torque_nm": float(i * 100), "gear_type": "spur"}
            for i in range(5)
        ]
        records = [
            _make_record("BigCorp", "company_meta", {"company_name": "BigCorp"}),
            _make_record("BigCorp", "products_gearbox", {
                "products_gearbox": products,
            }),
        ]
        builder = ConsolidatedReportBuilder(generator)
        data_sheet, _ = builder.gather(records, sample_schema, page_size=2)

        entity_cols = [c for c in data_sheet.columns if "products_gearbox" in c]
        assert len(entity_cols) == 3  # ceil(5/2) = 3 pages
        assert entity_cols[0] == "products_gearbox_list"
        assert entity_cols[1] == "products_gearbox_list_p2"
        assert entity_cols[2] == "products_gearbox_list_p3"

    def test_pagination_labels_show_ranges(self, generator, sample_schema):
        """Paginated columns get range labels like 'Products Gearbox (1-2)'."""
        products = [
            {"product_name": f"G{i}", "max_torque_nm": float(i), "gear_type": "spur"}
            for i in range(4)
        ]
        records = [
            _make_record("Corp", "company_meta", {"company_name": "Corp"}),
            _make_record("Corp", "products_gearbox", {"products_gearbox": products}),
        ]
        builder = ConsolidatedReportBuilder(generator)
        data_sheet, _ = builder.gather(records, sample_schema, page_size=2)

        assert "(1-2)" in data_sheet.labels["products_gearbox_list"]
        assert "(3-4)" in data_sheet.labels["products_gearbox_list_p2"]

    def test_pagination_distributes_items(self, generator, sample_schema):
        """Items are split across page columns."""
        products = [
            {"product_name": f"G{i}", "max_torque_nm": float(i), "gear_type": "spur"}
            for i in range(3)
        ]
        records = [
            _make_record("Corp", "company_meta", {"company_name": "Corp"}),
            _make_record("Corp", "products_gearbox", {"products_gearbox": products}),
        ]
        builder = ConsolidatedReportBuilder(generator)
        data_sheet, _ = builder.gather(records, sample_schema, page_size=2)

        row = data_sheet.rows[0]
        # Page 1: G0, G1
        assert "G0" in row["products_gearbox_list"]
        assert "G1" in row["products_gearbox_list"]
        assert "G2" not in row["products_gearbox_list"]
        # Page 2: G2
        assert "G2" in row["products_gearbox_list_p2"]

    def test_pagination_short_row_fills_na(self, generator, sample_schema):
        """Source_group with fewer entities than max gets N/A in overflow pages."""
        products_big = [
            {"product_name": f"G{i}", "max_torque_nm": float(i), "gear_type": "spur"}
            for i in range(4)
        ]
        records = [
            # BigCorp has 4 products
            _make_record("BigCorp", "company_meta", {"company_name": "BigCorp"}),
            _make_record("BigCorp", "products_gearbox", {"products_gearbox": products_big}),
            # SmallCorp has 1 product
            _make_record("SmallCorp", "company_meta", {"company_name": "SmallCorp"}),
            _make_record("SmallCorp", "products_gearbox", {
                "products_gearbox": [{"product_name": "Only", "max_torque_nm": 10.0, "gear_type": "spur"}],
            }),
        ]
        builder = ConsolidatedReportBuilder(generator)
        data_sheet, _ = builder.gather(records, sample_schema, page_size=2)

        # SmallCorp (sorted second) should have N/A in page 2
        small_row = data_sheet.rows[1]
        assert "Only" in small_row["products_gearbox_list"]
        assert small_row["products_gearbox_list_p2"] == "N/A"

    def test_pagination_provenance_per_page(self, generator, sample_schema):
        """Quality sheet computes per-page entity quality averages."""
        products = [
            {"product_name": f"G{i}", "max_torque_nm": float(i), "gear_type": "spur"}
            for i in range(4)
        ]
        entity_prov = [
            {"winning_weight": 0.9, "top_sources": ["s1"]},
            {"winning_weight": 0.8, "top_sources": ["s1"]},
            {"winning_weight": 0.7, "top_sources": ["s1"]},
            {"winning_weight": 0.6, "top_sources": ["s1"]},
        ]
        records = [
            _make_record("Corp", "company_meta", {"company_name": "Corp"}),
            _make_record("Corp", "products_gearbox", {
                "products_gearbox": products,
            }, provenance={
                "products_gearbox": {
                    "winning_weight": 0.75,
                    "top_sources": ["s1"],
                    "entity_provenance": entity_prov,
                },
            }),
        ]
        builder = ConsolidatedReportBuilder(generator)
        data_sheet, _ = builder.gather(records, sample_schema, page_size=2)

        records_by_sg = {"Corp": {rec.extraction_type: rec for rec in records}}
        quality, _ = build_provenance_sheets(data_sheet, records_by_sg, {"s1": "http://x"})

        # Page 1 entities: weights 0.9, 0.8 → avg 0.85
        assert quality.rows[0]["products_gearbox_list"] == pytest.approx(0.85)
        # Page 2 entities: weights 0.7, 0.6 → avg 0.65
        assert quality.rows[0]["products_gearbox_list_p2"] == pytest.approx(0.65)

    def test_provenance_key_map(self, generator, sample_schema):
        """Paginated columns map back to provenance key via provenance_key_map."""
        products = [
            {"product_name": f"G{i}", "max_torque_nm": float(i), "gear_type": "spur"}
            for i in range(4)
        ]
        records = [
            _make_record("Corp", "company_meta", {"company_name": "Corp"}),
            _make_record("Corp", "products_gearbox", {"products_gearbox": products}),
        ]
        builder = ConsolidatedReportBuilder(generator)
        data_sheet, _ = builder.gather(records, sample_schema, page_size=2)

        assert data_sheet.provenance_key_map["products_gearbox_list"] == "products_gearbox"
        assert data_sheet.provenance_key_map["products_gearbox_list_p2"] == "products_gearbox"


# ── TestThreeSheetExcel ──


class TestThreeSheetExcel:
    def test_always_three_sheets(self, generator, sample_schema, sample_records):
        builder = ConsolidatedReportBuilder(generator)
        data_sheet, _ = builder.gather(sample_records, sample_schema)

        # Build records_by_sg
        records_by_sg = {}
        for rec in sample_records:
            sg = rec.source_group
            if sg not in records_by_sg:
                records_by_sg[sg] = {}
            records_by_sg[sg][rec.extraction_type] = rec

        quality, sources = build_provenance_sheets(data_sheet, records_by_sg, {})

        formatter = ExcelFormatter()
        excel_bytes = formatter.create_multi_sheet_workbook([data_sheet, quality, sources])
        assert isinstance(excel_bytes, bytes)

        from io import BytesIO
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(excel_bytes))
        assert len(wb.sheetnames) == 3
        assert wb.sheetnames[0] == "Company Data"
        assert wb.sheetnames[1] == "Company Data - Quality"
        assert wb.sheetnames[2] == "Company Data - Sources"

    def test_quality_sheet_has_values(self, generator, sample_schema, sample_records):
        builder = ConsolidatedReportBuilder(generator)
        data_sheet, _ = builder.gather(sample_records, sample_schema)

        records_by_sg = {}
        for rec in sample_records:
            sg = rec.source_group
            if sg not in records_by_sg:
                records_by_sg[sg] = {}
            records_by_sg[sg][rec.extraction_type] = rec

        quality, _ = build_provenance_sheets(data_sheet, records_by_sg, {})
        # Acme has provenance for company_name
        acme_quality = quality.rows[0]
        assert acme_quality["company_name"] == pytest.approx(0.9)

    def test_entity_list_provenance_via_suffix_stripping(self, generator, sample_schema, sample_records):
        """Entity list column provenance resolves via _list suffix stripping."""
        builder = ConsolidatedReportBuilder(generator)
        data_sheet, _ = builder.gather(sample_records, sample_schema)

        records_by_sg = {}
        for rec in sample_records:
            sg = rec.source_group
            if sg not in records_by_sg:
                records_by_sg[sg] = {}
            records_by_sg[sg][rec.extraction_type] = rec

        source_url_map = {"s1": "https://example.com/a", "s2": "https://example.com/b"}
        quality, sources = build_provenance_sheets(data_sheet, records_by_sg, source_url_map)

        # products_gearbox_list should resolve via stripping _list → products_gearbox
        # Quality is average of per-entity winning_weights (0.9 + 0.8) / 2 = 0.85
        acme_quality = quality.rows[0]
        assert acme_quality["products_gearbox_list"] == pytest.approx(0.85)

        acme_sources = sources.rows[0]
        assert "https://example.com/a" in acme_sources["products_gearbox_list"]

    def test_sheet_names_sanitized(self):
        sheets = [
            SheetData(
                name="Invalid/Name:Here",
                rows=[{"a": 1}],
                columns=["a"],
                labels={"a": "A"},
            )
        ]
        formatter = ExcelFormatter()
        excel_bytes = formatter.create_multi_sheet_workbook(sheets)

        from io import BytesIO
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(excel_bytes))
        assert "/" not in wb.sheetnames[0]
        assert ":" not in wb.sheetnames[0]

    def test_existing_create_workbook_unchanged(self):
        """Ensure single-sheet create_workbook still works."""
        formatter = ExcelFormatter()
        rows = [{"name": "Alice", "age": 30}]
        columns = ["name", "age"]
        labels = {"name": "Name", "age": "Age"}
        result = formatter.create_workbook(rows, columns, labels, "Test")
        assert isinstance(result, bytes)
        assert len(result) > 0


# ── TestRenderMarkdown ──


class TestRenderMarkdown:
    def test_section_header(self, generator, sample_schema, sample_records):
        builder = ConsolidatedReportBuilder(generator)
        data_sheet, summary = builder.gather(sample_records, sample_schema)
        md = render_markdown([data_sheet], summary)
        assert "## Company Data" in md

    def test_no_hardcoded_title(self, generator, sample_schema, sample_records):
        """render_markdown should not inject its own H1 title."""
        builder = ConsolidatedReportBuilder(generator)
        data_sheet, summary = builder.gather(sample_records, sample_schema)
        md = render_markdown([data_sheet], summary)
        assert not md.startswith("# ")

    def test_total_count_in_summary(self, generator, sample_schema, sample_records):
        builder = ConsolidatedReportBuilder(generator)
        data_sheet, summary = builder.gather(sample_records, sample_schema)
        md = render_markdown([data_sheet], summary)
        assert "Total: 2" in md

    def test_pipe_escaping(self):
        sheets = [
            SheetData(
                name="Test",
                rows=[{"val": "has|pipe"}],
                columns=["val"],
                labels={"val": "Value"},
            )
        ]
        md = render_markdown(sheets, {"total_count": 1})
        assert "has\\|pipe" in md
        assert "has|pipe" not in md.split("\n")[-2]


# ── TestReportRequestModel ──


class TestReportRequestModel:
    def test_consolidated_valid(self):
        from models import ReportRequest, ReportType

        req = ReportRequest(
            type=ReportType.TABLE,
            group_by="consolidated",
        )
        assert req.group_by == "consolidated"

    def test_consolidated_requires_table(self):
        from models import ReportRequest, ReportType

        with pytest.raises(Exception):
            ReportRequest(
                type=ReportType.SINGLE,
                group_by="consolidated",
            )

    def test_removed_fields_absent(self):
        from models import ReportRequest, ReportType

        req = ReportRequest(type=ReportType.TABLE, group_by="consolidated")
        assert not hasattr(req, "layout")
        assert not hasattr(req, "entity_focus")
        assert not hasattr(req, "include_provenance")
        assert not hasattr(req, "provenance_sheets")
