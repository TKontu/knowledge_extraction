"""Tests for Excel formatter."""

from io import BytesIO

from openpyxl import load_workbook

from services.reports.excel_formatter import ExcelFormatter


class TestExcelFormatter:
    def test_create_workbook_basic(self):
        """Test basic workbook creation."""
        formatter = ExcelFormatter()
        rows = [
            {"name": "Company A", "has_feature": True, "count": 100},
            {"name": "Company B", "has_feature": False, "count": 50},
        ]
        columns = ["name", "has_feature", "count"]

        result = formatter.create_workbook(rows, columns)

        assert isinstance(result, bytes)
        wb = load_workbook(BytesIO(result))
        ws = wb.active
        assert ws.cell(1, 1).value == "Name"
        assert ws.cell(2, 1).value == "Company A"
        assert ws.cell(2, 2).value == "Yes"
        assert ws.cell(3, 2).value == "No"

    def test_format_value_handles_types(self):
        """Test value formatting for different types."""
        formatter = ExcelFormatter()

        assert formatter._format_value(None) == "N/A"
        assert formatter._format_value(True) == "Yes"
        assert formatter._format_value(False) == "No"
        assert formatter._format_value(["a", "b"]) == "a\nb"
        assert formatter._format_value("text") == "text"

    def test_humanize(self):
        """Test field name humanization."""
        formatter = ExcelFormatter()

        assert formatter._humanize("field_name") == "Field Name"
        assert formatter._humanize("manufactures_gearboxes") == "Manufactures Gearboxes"

    def test_column_width_uses_longest_line(self):
        """Test column width is based on longest line, not total string length."""
        formatter = ExcelFormatter()
        # Value with multiple lines - longest line is 19 chars
        rows = [{"col": "short\nvery_long_line_here\nmed"}]
        columns = ["col"]

        result = formatter.create_workbook(rows, columns)

        wb = load_workbook(BytesIO(result))
        ws = wb.active
        width = ws.column_dimensions["A"].width

        # Width should be based on longest line (19) + padding (2) = 21
        # Not total length (29) + padding (2) = 31
        assert width <= 25, f"Width {width} should be based on longest line (~21), not total length (~31)"
