"""Excel report generation using openpyxl."""

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExcelFormatter:
    """Formats tabular data into Excel workbooks."""

    def __init__(self) -> None:
        self._header_font = Font(bold=True, color="FFFFFF")
        self._header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        self._border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def create_workbook(
        self,
        rows: list[dict[str, Any]],
        columns: list[str],
        column_labels: dict[str, str] | None = None,
        sheet_name: str = "Report",
    ) -> bytes:
        """Create Excel workbook from row data.

        Args:
            rows: List of dicts, each representing a row
            columns: Column field names in order
            column_labels: Optional mapping of field names to display labels
            sheet_name: Name for the worksheet

        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        ws = wb.active
        # Sanitize sheet name - Excel doesn't allow: \ / ? * [ ] :
        # and max length is 31 characters
        safe_name = sheet_name
        for char in r"\/?*[]:":
            safe_name = safe_name.replace(char, "-")
        ws.title = safe_name[:31]

        labels = column_labels or {}

        # Write headers
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = labels.get(col_name, self._humanize(col_name))
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = self._border

        # Write data rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = row_data.get(col_name)
                cell.value = self._format_value(value)
                cell.border = self._border
                cell.alignment = Alignment(wrap_text=True)

        # Auto-adjust column widths
        for col_idx, col_name in enumerate(columns, 1):
            max_length = len(labels.get(col_name, self._humanize(col_name)))
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max_length + 2, 50
            )

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    def _humanize(self, field_name: str) -> str:
        """Convert field_name to Human Readable Label."""
        return field_name.replace("_", " ").title()

    def _format_value(self, value: Any) -> str:
        """Format extraction values for display."""
        if value is None:
            return "N/A"
        if isinstance(value, bool):
            return "Yes" if value else "No"
        if isinstance(value, list):
            return ", ".join(str(v) for v in value)
        return str(value)
